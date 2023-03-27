"""
This file defines the process for calculating the Ford function.
Only used for Ford project.

Note: Spot and DBL calculations are separate from each other.

@Author: Siwei.Lu
@Date: 2022.11.27
"""

import openpyxl
from scipy import stats
import matplotlib.pyplot as plt

from Core.project.ford.dbl_calc import *
from Core.project.ford.spot_calc import *
from Core.dataset_handler.ford_dataset_handler import FordDatasetHandler
from Core.can_related.canoe_connect import CanoeSync
from Core.project.ford.msg_sig_env_def import Message, Signal
from Core.project.ford.cdd_qualifier_def import SubFunctions


class CalHandler:
    def __init__(self, dataset: FordDatasetHandler, app: CanoeSync, battery):
        self.__dataset = dataset
        self.__battery = battery
        self.__app = app
        self.__side = "Left ECU"

        self.__excel_data = self.__dataset.excel_data

        self._get_used_para_in_dataset()
        self._get_traffic_style()

    @staticmethod
    def get_excel_data_test(self, path="D:/001_Ford/005_Dataset/12B/KBD_LDM12B_VB009_20220930.xlsm"):
        sheet_data_dict = {}
        data = openpyxl.load_workbook(path)
        data.close()
        for key, value in UsedSheet.__members__.items():
            sheet = value.value
            if sheet not in data.sheetnames:
                raise Exception("The sheet named {} is not exist in the file {}".format(sheet, path))
            sheet_data_dict[sheet] = data.get_sheet_by_name(sheet)
        test = {
            "CH_4":
                [
                    {"Chip_1":
                         {"position": (1, 0, 1, 0),
                          "intensity":
                              {"Left": [{"CLASS_C", 16, "CLASS_E", 10}],
                               "Right": [{"CLASS_C", 16, "CLASS_E", 10}]
                               }
                          }
                     }
                ]
        }
        my_list = []
        for index in range(4, 29):
            my_list.append(sheet_data_dict['Pixel-System'].cell(row=10, column=index).value)
        test["CH_4"][0]["Chip_1"]["position"] = tuple(my_list[4:8])
        print(test)

    @property
    def side(self):
        return self.__side

    @side.setter
    def side(self, side):
        self.__side = side

    def _get_used_para_in_dataset(self):
        # Pixel
        self.__pixel_site_value = self.__dataset.pixel_site_value
        self.__sec01_site_value = self.__dataset.sec01_site_value
        self.__sec23_site_value = self.__dataset.sec23_site_value
        self.__sec45_site_value = self.__dataset.sec45_site_value

        self.__matrix_add_swi = self.__dataset.matrix_add_swi

        self.__sec01_matrix_add_swi = self.__dataset.sec01_matrix_add_swi
        self.__sec23_matrix_add_swi = self.__dataset.sec23_matrix_add_swi
        self.__sec45_matrix_add_swi = self.__dataset.sec45_matrix_add_swi

        # GFHB
        self.__GFHB_para = self.__dataset.GFHB_para  # GFHB function列的参数
        self.__HLT_AnDistLutBrkPnts = self.__dataset.HLT_AnDistLutBrkPnts

        # DBL
        self.__DBL_shiftAway = self.__dataset.DBL_shiftAway
        self.__DBL_leading_fac = self.__dataset.DBL_leading_fac
        self.__DBL_follow_fac = self.__dataset.DBL_follow_fac
        self.__DBL_speed = self.__dataset.DBL_speed

        self.__DBL_base_row = self.__dataset.DBL_base_row
        self.__DBL_row1_enable = self.__dataset.DBL_row1_enable
        self.__DBL_row2_enable = self.__dataset.DBL_row2_enable
        self.__DBL_row3_enable = self.__dataset.DBL_row3_enable

        # calc
        self.__sec01_notNull_index = get_not_null_index(part_reverse(self.__sec01_site_value[0], 16))
        self.__sec23_notNull_index = get_not_null_index(part_reverse(self.__sec23_site_value[0], 16))
        self.__sec45_notNull_index = get_not_null_index(part_reverse(self.__sec45_site_value[0], 16))

        self.__ch_function = self.__dataset.ch_function
        self.__ch_fun_pixel_dict = self._get_channel_function_dict(FunMapToENV.Pixel.name, 0, 2)

        # calculate
        # self.__spot_area = None
        # self.__smooth_area = None
        # self.__spot_smooth_HB_intensity = None
        # self.__pixel_int_after_spot = None
        # self.__dbl_move_ang = 0
        # self.__pixel_final_int = None

    def _get_used_can_signal_status(self):
        # todo:move to class:can
        self.__traffic_style = self.__app.get_signal(Message.AFSRq.value, Signal.Traffic_Style_Rq.value)

        if self.__side == ecu_side.left.value:
            class_level_panel = self.__app.get_signal(Message.AFSRq.value, Signal.LowBeamLightClassL_D_Rq.value)
            self.__dbl_move_ang = round(float(self.__app.get_signal(Message.AFSRq.value,
                                                                    Signal.SwvlTrgtLeft_An_Rq.value)), 2)
        else:
            class_level_panel = self.__app.get_signal(Message.AFSRq.value, Signal.LowBeamLightClassR_D_Rq.value)
            self.__dbl_move_ang = round(float(self.__app.get_signal(Message.AFSRq.value,
                                                                    Signal.SwvlTrgtRight_An_Rq.value)), 2)

    def _get_beam_variant(self):
        stream = self.__app.send_diag_req(SubFunctions.Method_2_Data_Type_1_Read.value)
        self.__mode = stream[4] + 1

    def _get_class_mod_intensity(self):
        self._get_beam_variant()

        mode_start_row = self.__mode * 100 - 98  # dataset development sheet
        mode_end_row = self.__mode * 100 - 3  # dataset development sheet

        self.__flb_intensity = self.__dataset.get_row_column_array(UsedSheet.development.value,
                                                                   [self.__mode * 100 - 99,
                                                                    self.__mode * 100 - 99,
                                                                    1,
                                                                    19])

        self.__class_intensity = self.__dataset.get_row_column_array(UsedSheet.development.value,
                                                                     [mode_start_row, mode_end_row, 1, 19])

    def _get_dbl_mode_para(self):
        self._get_beam_variant()

        start_col = (self.__mode - 1) * 19 + 4
        end_col = (self.__mode - 1) * 19 + 39

        self.__DBL_move_para = self.__dataset.get_row_column_array(UsedSheet.dbl.value,
                                                                   [22, 39, start_col, end_col])

        self.__sec01_intensity = self.__dataset.get_row_column_array(UsedSheet.development.value,
                                                                     [self.__mode * 100 - 98,
                                                                      self.__mode * 100 - 67,
                                                                      1,
                                                                      19])
        self.__sec23_intensity = self.__dataset.get_row_column_array(UsedSheet.development.value,
                                                                     [self.__mode * 100 - 66,
                                                                      self.__mode * 100 - 35,
                                                                      1,
                                                                      19])
        self.__sec45_intensity = self.__dataset.get_row_column_array(UsedSheet.development.value,
                                                                     [self.__mode * 100 - 34,
                                                                      self.__mode * 100 - 3,
                                                                      1,
                                                                      19])

    def _get_channel_function_dict(self, fun, key, value):
        """
        Get channel,function,MatrixChip dict
        :param fun: function configured in dataset
        :param key: 0: CH_x, 1: Function, 2: Matrix Chip
        :param value: 0: CH_x, 1: Function, 2: Matrix Chip
        :return: dict
        """
        channel_function_dict = {}

        for i in range(len(self.__ch_function[1])):
            if self.__ch_function[1][i].replace(" ", "_") == fun:
                channel_function_dict[self.__ch_function[key][i]] = self.__ch_function[value][i]

        # print(channel_function_dict)
        return channel_function_dict

    def _get_traffic_style(self):
        self.__traffic_style = self.__app.get_signal(Message.AFSRq.value, Signal.Traffic_Style_Rq.value)

    def _get_ecu_voltage(self):
        if self.__side == ecu_side.left.value:
            data = self.__app.send_diag_req(SubFunctions.Left_Headlamp_Power_Supply_Read.value)
        else:
            data = self.__app.send_diag_req(SubFunctions.Right_Headlamp_Power_Supply_Read.value)

        self.__ecu_voltage = round(data[3] / 4, 2)

    def _get_class_level(self):  # 获取左右ECU的class level
        if self.__side == ecu_side.left.value:
            class_level_panel = self.__app.get_signal(Message.AFSRq.value, Signal.LowBeamLightClassL_D_Rq.value)
        else:
            class_level_panel = self.__app.get_signal(Message.AFSRq.value, Signal.LowBeamLightClassR_D_Rq.value)

        # class_hb = self.__app.get_signal(Message.HighBeamMaster.value, Signal.HeadLghtHi_D_Rq.value)  # 12B
        class_hb = self.__app.get_signal(Message.HiBeamRq.value, Signal.HBHL_Beam_D_Rq.value)  # 12C

        if int(float(class_level_panel)) in {0, 5, 6, 7, 8, 14, 15}:
            self.__class_level = 0  # Class C
        if int(float(class_level_panel)) in {1, 2, 3, 4}:
            self.__class_level = 3  # v1
        if int(float(class_level_panel)) in {9}:
            self.__class_level = 1  # E
        if int(float(class_level_panel)) in {10}:
            self.__class_level = 4  # E1
        if int(float(class_level_panel)) in {11}:
            self.__class_level = 5  # E2
        if int(float(class_level_panel)) in {12}:
            self.__class_level = 6  # E3
        if int(float(class_level_panel)) in {13}:
            self.__class_level = 2  # W
        self.__classLB_level = self.__class_level
        self.__dbl_move_class_level = self.__class_level

        # If HB is requested
        if int(float(class_hb)) in {1, 2, 3}:
            self.__class_level = 7
            self.__dbl_move_class_level = 7

        if int(float(self.__traffic_style)) == 1:
            self.__classLB_level = self.__classLB_level + 10
            self.__class_level = self.__class_level + 10

    def get_flb_class_intensity(self):
        self._get_traffic_style()
        self._get_class_level()
        self._get_class_mod_intensity()
        self._get_ecu_voltage()

        ch_flb_dict = self._get_channel_function_dict(FunMapToENV.Foreground_LB.name, 0, 1)
        flb_int = self.__flb_intensity[self.__class_level]

        if float(self.__ecu_voltage) < 9:
            flb_int = self.__flb_intensity[0]

        return ch_flb_dict, flb_int

    def get_pixel_class_intensity(self):
        self._get_traffic_style()
        self._get_class_level()
        self._get_class_mod_intensity()
        self._get_ecu_voltage()

        self.__pixel_final_int = self.__class_intensity[self.__class_level]

        if float(self.__ecu_voltage) < 9:
            self.__pixel_final_int = self.__class_intensity[0]

        self._check_pixel_min_pwm()

    # ***********************  Spot  ***********************
    def _get_adjust_spot_ang(self, spot_num1, spot_num2):
        HLT_AnLuBrkPnts = self.__dataset.HLT_AnLuBrkPnts
        adjust_left_ang, adjust_right_ang, adjust_top_ang, adjust_bot_ang = [], [], [], []

        # Get the left/right/top/bottom angle from can signal
        for i in range(spot_num1, spot_num2 + 1):
            if self.__side == ecu_side.right.value:
                left_ang = float(self.__app.get_signal(Message.HighBeamMaster.value,
                                                       "Hbm%sRightLghtLeft_An_Rq" % str(i)))
                right_ang = float(self.__app.get_signal(Message.HighBeamMaster.value,
                                                        "Hbm%sRightLghtRigh_An_Rq" % str(i)))
            elif self.__side == ecu_side.left.value:
                left_ang = float(self.__app.get_signal(Message.HighBeamMaster.value,
                                                       "Hbm%sLeftLghtLeft_An_Rq" % str(i)))
                right_ang = float(self.__app.get_signal(Message.HighBeamMaster.value,
                                                        "Hbm%sLeftLghtRight_An_Rq" % str(i)))
            else:
                left_ang = 0
                right_ang = 0

            top_ang = float(self.__app.get_signal(Message.HighBeamMaster.value,
                                                  "Hbm%sBrdrTop_An_Rq" % str(i)))
            bot_ang = float(self.__app.get_signal(Message.HighBeamMaster.value,
                                                  "Hbm%sBrdrBottom_An_Rq" % str(i)))

            left_ang = signal_value_out_of_range(left_ang, HLT_AnLuBrkPnts[0])
            right_ang = signal_value_out_of_range(right_ang, HLT_AnLuBrkPnts[2])
            top_ang = signal_value_out_of_range(top_ang, HLT_AnLuBrkPnts[4])
            bot_ang = signal_value_out_of_range(bot_ang, HLT_AnLuBrkPnts[6])

            left_offset = get_var_in_array_site(HLT_AnLuBrkPnts[0], left_ang, 0)
            right_offset = get_var_in_array_site(HLT_AnLuBrkPnts[2], right_ang)
            top_offset = get_var_in_array_site(HLT_AnLuBrkPnts[4], top_ang)
            bot_offset = get_var_in_array_site(HLT_AnLuBrkPnts[6], bot_ang)

            # 计算斜率
            slope_left, intercept_left, r, p, s = stats.linregress(HLT_AnLuBrkPnts[0][left_offset:left_offset + 2],
                                                                   HLT_AnLuBrkPnts[1][left_offset:left_offset + 2])
            slope_right, intercept_right, r, p, s = stats.linregress(HLT_AnLuBrkPnts[2][right_offset:right_offset + 2],
                                                                     HLT_AnLuBrkPnts[3][right_offset:right_offset + 2])
            slope_top, intercept_top, r, p, s = stats.linregress(HLT_AnLuBrkPnts[4][top_offset:top_offset + 2],
                                                                 HLT_AnLuBrkPnts[5][top_offset:top_offset + 2])
            slope_bot, intercept_bot, r, p, s = stats.linregress(HLT_AnLuBrkPnts[6][bot_offset:bot_offset + 2],
                                                                 HLT_AnLuBrkPnts[7][bot_offset:bot_offset + 2])

            # 线性变化
            adjust_left_ang.append(round(left_ang + slope_left * left_ang + intercept_left, 4))
            adjust_right_ang.append(round(right_ang + slope_right * right_ang + intercept_right, 4))
            adjust_top_ang.append(round(top_ang + slope_top * top_ang + intercept_top, 4))
            adjust_bot_ang.append(round(bot_ang + slope_bot * bot_ang + intercept_bot, 4))

            if adjust_left_ang[i - 1] <= (adjust_right_ang[i - 1] + 0.2):
                adjust_right_ang[i - 1] = round(adjust_left_ang[i - 1] - 0.2, 4)
            if adjust_top_ang[i - 1] >= (adjust_bot_ang[i - 1] - 0.2):
                adjust_top_ang[i - 1] = round(adjust_bot_ang[i - 1] - 0.2, 4)

        self.__spot_area = [adjust_left_ang,
                            adjust_right_ang,
                            adjust_top_ang,
                            adjust_bot_ang]

    def _get_smooth_area(self):
        m3 = self.__HLT_AnDistLutBrkPnts[0][10]  # M3_HLT_SnDistLutBrkPnts(Deg)
        smooth_area_left, smooth_area_right, smooth_area_up, smooth_area_bot = [], [], [], []

        for i in range(1, 7):
            left_smooth_fac = float(
                self.__app.get_signal(Message.HighBeamMaster.value, ("Hbm%sTrnstnLeft_An_Rq" % str(i))))
            right_smooth_fac = float(
                self.__app.get_signal(Message.HighBeamMaster.value, ("Hbm%sTrnstnRight_An_Rq" % str(i))))

            smooth_area_left.append((left_smooth_fac * m3 / 3) + self.__spot_area[0][i - 1])
            smooth_area_right.append(-(right_smooth_fac * m3 / 3) + self.__spot_area[1][i - 1])
            smooth_area_up.append(self.__spot_area[2][i - 1])
            smooth_area_bot.append(self.__spot_area[3][i - 1])

        self.__smooth_area = [smooth_area_left, smooth_area_right,
                              smooth_area_up, smooth_area_bot]

    def get_status_after_spot(self):
        self._get_traffic_style()
        self._get_class_level()
        self._get_class_mod_intensity()

        # get adjust spot ang
        self._get_adjust_spot_ang(1, 6)
        # get smooth area
        self._get_smooth_area()

        for i in range(1, 7):
            self._get_pixel_area_intensity_after_spot(i)

        self._check_pixel_min_pwm()

    @staticmethod
    def _add_site(ax1, left, right, up, bot, color, al):
        ax1.add_patch(
            plt.Rectangle(
                (float(left), float(up)),
                float(right) - float(left),
                float(bot) - float(up),
                color=color,
                alpha=al
            )
        )

    def get_pixel_site_sec(self, function):
        plt.close()
        fig1 = plt.figure()
        max_left = min_right = min_up = max_down = 0
        ax1 = fig1.add_subplot(111, aspect='equal')

        #  pixel range
        for i in range(len(self.__pixel_site_value[0])):
            if self.__pixel_site_value[0][i] != 'NU':
                # pixel site
                self._add_site(ax1, self.__pixel_site_value[0][i], self.__pixel_site_value[1][i],
                               self.__pixel_site_value[2][i], self.__pixel_site_value[3][i], 'blue', 0.1)

                add_pixel_intensity(ax1,
                                    self.__pixel_site_value[0][i],
                                    self.__pixel_site_value[3][i],
                                    self.__pixel_final_int[i],
                                    9)

                if self.__pixel_site_value[0][i] > max_left:
                    max_left = self.__pixel_site_value[0][i]
                if self.__pixel_site_value[1][i] < min_right:
                    min_right = self.__pixel_site_value[1][i]
                if self.__pixel_site_value[2][i] < min_up:
                    min_up = self.__pixel_site_value[2][i]
                if self.__pixel_site_value[3][i] > max_down:
                    max_down = self.__pixel_site_value[3][i]

        # spot range and smooth range
        if function is Function.Spot.value:
            for i in range(0, 6):
                if int(float(self.__app.get_signal(Message.HighBeamMaster.value, ("Hbm%sMde_D_Rq" % str(i + 1))))) != 0:
                    self._add_site(ax1, self.__spot_area[0][i], self.__spot_area[1][i], self.__spot_area[2][i],
                                   self.__spot_area[3][i], "red", 0.3)
                    self._add_site(ax1, self.__smooth_area[0][i], self.__spot_area[0][i], self.__spot_area[2][i],
                                   self.__spot_area[3][i], "yellow", 0.3)
                    self._add_site(ax1, self.__spot_area[1][i], self.__smooth_area[1][i], self.__spot_area[2][i],
                                   self.__spot_area[3][i], "yellow", 0.3)

        plt.xlim(max_left + 2, min_right - 2)
        plt.ylim(min_up - 2, max_down + 2)
        ax1.invert_yaxis()
        plt.show()

    def get_pixel_ch_int_dict(self):
        tmp_dict = {}

        for key, value in self.__ch_fun_pixel_dict.items():  # key: CH-x,   value:  Matrix Chip
            int_dict = {}

            for i in range(len(self.__matrix_add_swi[0])):  # pixel:  Matrix address, Bypass switch
                if self.__matrix_add_swi[0][i] == value:  # chip
                    if self.__pixel_final_int[i] is not None:
                        int_dict.update({self.__matrix_add_swi[1][i]: self.__pixel_final_int[i]})
                    else:
                        # Set intensity to 0 if intensity is None
                        int_dict.update({self.__matrix_add_swi[1][i]: 0})

            for switch in range(1, 13):
                if switch not in int_dict.keys():
                    int_dict.update({switch: 0})

            tmp_dict.update({key: int_dict})

        return tmp_dict

    def _get_pixel_area_intensity_after_spot(self, spot_num):
        self.__pixel_int_after_spot = self.__class_intensity[self.__class_level].copy()

        spot_act_sta = int(
            float(self.__app.get_signal(Message.HighBeamMaster.value, ("Hbm%sMde_D_Rq" % str(spot_num)))))
        spot_num = spot_num - 1  # spot1~6 to 0~5

        if spot_act_sta == 0:
            print("Spot{}: No spot".format(spot_num + 1))
        else:
            # For Left smooth area and Spot Area
            for i in range(len(self.__pixel_site_value[0])):
                if self.__pixel_site_value[0][i] != 'NU':
                    dist = self.__pixel_site_value[1][i] - self.__spot_area[0][spot_num]  # 用LED的右角度和spot左角度
                    self._get_left_smooth_intensity(spot_num, i, dist, spot_act_sta)
                    self._get_spot_area_intensity(spot_num, i, spot_act_sta)
            # For right smooth area
            for i in range(len(self.__pixel_site_value[1])):
                if self.__pixel_site_value[1][i] != 'NU':
                    dist = self.__spot_area[1][spot_num] - self.__pixel_site_value[0][i]  # 用LED的左角度和spot右角度
                    self._get_right_smooth_intensity(spot_num, i, dist, spot_act_sta)
                    # self.__get_spot_area_intensity(spot_num, i, spot_act_sta)

        self.__pixel_final_int = self.__pixel_int_after_spot

    def _get_spot_area_intensity(self, spot_num, i, spot_act_sta):
        """
        获取spot区域灯的亮度，pixelLB需要和classLB intensity比较
        :param spot_num: spot1~6
        :param i:
        :param spot_act_sta: 1:DarkSpot  2:DeglareSpot
        :return:
        """
        if (((self.__pixel_site_value[0][i] <= self.__spot_area[0][spot_num]) and
             (self.__pixel_site_value[0][i] >= self.__spot_area[1][spot_num])) or
            ((self.__pixel_site_value[1][i] <= self.__spot_area[0][spot_num]) and
             (self.__pixel_site_value[1][i] >= self.__spot_area[1][spot_num]))) and \
                (((self.__pixel_site_value[2][i] <= self.__spot_area[3][spot_num]) and
                  (self.__pixel_site_value[2][i] >= self.__spot_area[2][spot_num])) or
                 ((self.__pixel_site_value[3][i] <= self.__spot_area[3][spot_num]) and
                  (self.__pixel_site_value[3][i] >= self.__spot_area[2][spot_num])) or
                 ((self.__spot_area[3][spot_num] <= self.__pixel_site_value[3][i]) and
                  (self.__spot_area[3][spot_num] >= self.__pixel_site_value[2][i]))):
            if i <= 63:  # class LB 64行
                spot_int = self._cal_spot_int(spot_act_sta, i)
                self.__pixel_int_after_spot[i] = self._spot_kink_strategy_for_spot(spot_int, i, spot_act_sta)
            else:
                if spot_act_sta == 1:  # Dark Spot
                    self.__pixel_int_after_spot[i] = 0
                else:  # Deglare Spot
                    self.__pixel_int_after_spot[i] = self._cal_spot_int(spot_act_sta, i)

    def _get_right_smooth_intensity(self, spot_num, i, dist, spot_act_sta):
        if (self.__pixel_site_value[0][i] >= self.__smooth_area[1][spot_num]) and \
                (self.__pixel_site_value[0][i] < self.__spot_area[1][spot_num]) and \
                (((self.__pixel_site_value[2][i] <= self.__spot_area[3][spot_num]) and
                  (self.__pixel_site_value[2][i] >= self.__spot_area[2][spot_num])) or
                 ((self.__pixel_site_value[3][i] <= self.__spot_area[3][spot_num]) and
                  (self.__pixel_site_value[3][i] >= self.__spot_area[2][spot_num])) or
                 ((self.__spot_area[3][spot_num] <= self.__pixel_site_value[3][i]) and
                  (self.__spot_area[3][spot_num] >= self.__pixel_site_value[2][i]))):
            HLT_AnDisLuIntensity = self._cal_AnDisLuIntensity(dist, 1, spot_act_sta, spot_num)
            smooth_intensity = round(HLT_AnDisLuIntensity * self.__class_intensity[self.__class_level][i]) / 100

            if i <= 63:
                smooth_area_int = self._spot_kink_strategy_for_smooth(smooth_intensity, i, spot_act_sta)
            else:
                smooth_area_int = smooth_intensity
            self.__pixel_int_after_spot[i] = smooth_area_int

            # print("spot{} smooth_intensity_R = {}%".format(spot_num + 1, smooth_area_int))

    def _get_left_smooth_intensity(self, spot_num, i, dist, spot_act_sta):
        if (self.__pixel_site_value[1][i] <= self.__smooth_area[0][spot_num]) and \
                (self.__pixel_site_value[1][i] > self.__spot_area[0][spot_num]) and \
                (((self.__pixel_site_value[2][i] <= self.__spot_area[3][spot_num]) and
                  (self.__pixel_site_value[2][i] >= self.__spot_area[2][spot_num])) or
                 ((self.__pixel_site_value[3][i] <= self.__spot_area[3][spot_num]) and
                  (self.__pixel_site_value[3][i] >= self.__spot_area[2][spot_num])) or
                 ((self.__spot_area[3][spot_num] <= self.__pixel_site_value[3][i]) and
                  (self.__spot_area[3][spot_num] >= self.__pixel_site_value[2][i]))):
            HLT_AnDisLuIntensity = self._cal_AnDisLuIntensity(dist, 0, spot_act_sta, spot_num)  # 0：left smooth
            smooth_intensity = round(HLT_AnDisLuIntensity * self.__class_intensity[self.__class_level][i]) / 100
            if i < 63:
                smooth_area_int = self._spot_kink_strategy_for_smooth(smooth_intensity, i, spot_act_sta)
            else:
                smooth_area_int = smooth_intensity

            self.__pixel_int_after_spot[i] = smooth_area_int

            # print("spot{} smooth_intensity_L = {}% i={}".format(spot_num + 1, smooth_area_int, i))

    def _spot_kink_strategy_for_spot(self, spot_int, i, spot_act_sta):
        """
        Interaction between Spot and LB_For spot: SYSRS_28893
        :param spot_int:
        :param i:
        :param spot_act_sta:
        :return:
        """
        spotkinkstratery = self.__GFHB_para[8]
        spotkindivvalue = self.__GFHB_para[9]

        spot_LB_area_int = self._spot_LB_int_compare(spot_int, i)

        # DarkSpot才生效
        if (spotkinkstratery == SpotKinkStrategy.TurnOff.value) and \
                (spot_act_sta == 1):
            spot_LB_area_int = 0
        elif spotkinkstratery == SpotKinkStrategy.PartialTurnOff.value and \
                (spot_act_sta == 1):
            if int(float(self.__traffic_style)) == 1:  # Right hand
                if self.__pixel_site_value[0][i] >= spotkindivvalue:  # Strategy2
                    spot_LB_area_int = 0
            if int(float(self.__traffic_style)) == 0:  # Left hand
                if self.__pixel_site_value[1][i] <= -spotkindivvalue:  # Strategy2
                    spot_LB_area_int = 0

        return spot_LB_area_int

    def _spot_kink_strategy_for_smooth(self, smooth_intensity, i, spot_act_sta):
        """
        Interaction between Spot and LB_For smooth: SYSRS_28893
        :param smooth_intensity:
        :param i:
        :param spot_act_sta:
        :return:
        """
        spotkinkstratery = self.__GFHB_para[8]
        spotkindivvalue = self.__GFHB_para[9]

        smooth_area_int = self._spot_LB_int_compare(smooth_intensity, i)  # Strategy1

        if spotkinkstratery == SpotKinkStrategy.TurnOff.value and \
                (spot_act_sta == 1):  # Strategy2
            smooth_area_int = smooth_intensity
        elif spotkinkstratery == SpotKinkStrategy.PartialTurnOff.value and \
                (spot_act_sta == 1):
            if int(float(self.__traffic_style)) == 1:  # Right hand
                if self.__pixel_site_value[0][i] >= spotkindivvalue:  # Strategy3--2
                    smooth_area_int = smooth_intensity
            if int(float(self.__traffic_style)) == 0:  # Left hand
                if self.__pixel_site_value[1][i] <= -spotkindivvalue:  # Strategy3--2
                    smooth_area_int = smooth_intensity

        return smooth_area_int

    def _cal_AnDisLuIntensity(self, dist, direction, spot_act_sta, spot_num):
        if direction == 1:  # 1：right smooth
            smooth_fac = self.__app.get_signal(Message.HighBeamMaster.value,
                                               ("Hbm%sTrnstnRight_An_Rq" % str(spot_num + 1)))
        elif direction == 0:  # 0：left smooth
            smooth_fac = self.__app.get_signal(Message.HighBeamMaster.value,
                                               ("Hbm%sTrnstnLeft_An_Rq" % str(spot_num + 1)))
        else:
            raise Exception("direction = {}, not in [0, 1]".format(direction))

        dist = dist * 3 / float(smooth_fac)
        offset = get_var_in_array_site(self.__HLT_AnDistLutBrkPnts[0], dist)
        slope, intercept, r, p, std = stats.linregress(self.__HLT_AnDistLutBrkPnts[0][offset:offset + 2],
                                                       self.__HLT_AnDistLutBrkPnts[1][offset:offset + 2])
        if spot_act_sta == 1:
            HLT_AnDisLuIntensity = slope * dist + intercept  # DarkSpot
        elif spot_act_sta == 2:
            HLT_AnDisLuIntensity = slope * dist + intercept + self.__GFHB_para[1]  # Deglare Spot
            if HLT_AnDisLuIntensity >= 100:
                HLT_AnDisLuIntensity = 100
        else:
            raise Exception("spot_act_sta = {}, not in [1, 2]".format(spot_act_sta))

        return HLT_AnDisLuIntensity

    def _cal_spot_int(self, spot_act_sta, i):
        if spot_act_sta == 1:
            spot_int = 0
        elif spot_act_sta == 2:
            # reduce intensity
            spot_int = round(self.__GFHB_para[1] * (self.__class_intensity[self.__class_level][i])) / 100
        else:
            raise Exception("__cal_spot_int Error: spot_act_sta is {}".format(spot_act_sta))
        # print("spot_int = {}, i = {}".format(spot_int, i))

        return spot_int

    def _spot_LB_int_compare(self, spot_int, i):
        classLB_int = float(self.__class_intensity[self.__classLB_level][i])
        if classLB_int > spot_int:
            spot_area_int = classLB_int
        else:
            spot_area_int = spot_int

        return spot_area_int

    # ***********************  DBL  ***********************
    def get_status_after_dbl(self):
        self._get_traffic_style()
        self._get_class_level()
        self._get_dbl_mode_para()
        self._get_dbl_move_ang()

        self._cal_int_after_dbl_move(self.__dbl_move_ang)

    def get_dbl_speed(self, enter_delta_ang, enter_delta_intensity, label_dbl_time):
        if not (enter_delta_ang.text() or enter_delta_intensity.text()):
            label_dbl_time.setText("↑↑↑ Please Enter DBL Angle and DBL intensity ↑↑↑")
        else:
            # print("enter_deltaAng = ", enter_deltaAng.text())
            unit = '%/s'
            changeRate = float(enter_delta_ang.text()) * 1000 / 30

            for i in range(len(self.__DBL_speed)):
                if changeRate > self.__DBL_speed[49]:
                    spd = len(self.__DBL_speed)
                elif changeRate >= self.__DBL_speed[i]:
                    spd = i + 3
                    if changeRate >= (self.__DBL_speed[i] + self.__DBL_speed[i + 1]) / 2:
                        spd = i + 4
                else:
                    spd = 3
                    print("spd = ", spd)

            base_speed = float(self.__dataset.get_row_column_array(UsedSheet.dbl.value, [spd, spd, 102, 102]))
            leading_speed = float(self.__DBL_leading_fac * base_speed)
            follow_speed = float(self.__DBL_follow_fac * base_speed)

            base_time = float(enter_delta_intensity.text()) / base_speed
            leading_time = float(enter_delta_intensity.text()) / leading_speed
            following_time = float(enter_delta_intensity.text()) / follow_speed

            if changeRate > self.__DBL_speed[0][49]:
                base_speed = 0
                leading_speed = 0
                follow_speed = 0
                base_time = 0
                leading_time = 0
                following_time = 0
            if base_time > 10.2:
                base_time = 10.2
            if leading_time > 10.2:
                leading_time = 10.2
            if following_time > 10.2:
                following_time = 10.2

            label_dbl_time.setText("base_speed      = %.2f %s,    base LED fade time      = %.2f s \r\n"
                                   "leading_speed   = %.2f %s,    leading LED fade time   = %.2f s \r\n"
                                   "following_speed = %.2f %s,    following LED fade time = %.2f s \r\n"
                                   % (base_speed, unit, base_time,
                                      leading_speed, unit, leading_time,
                                      follow_speed, unit, following_time)
                                   )

    def _get_dbl_move_ang(self):
        """
        This function is used to calculate the actual DBL move angle of ROW1/2/3.
        """
        if self.__side == ecu_side.right.value:
            self.__dbl_move_ang = round(float(self.__app.get_signal(Message.AFSRq.value,
                                                                    Signal.SwvlTrgtRight_An_Rq.value)), 2)
        elif self.__side == ecu_side.left.value:
            self.__dbl_move_ang = round(float(self.__app.get_signal(Message.AFSRq.value,
                                                                    Signal.SwvlTrgtLeft_An_Rq.value)), 2)
        else:
            raise Exception("ECU side not correct! {}".format(self.__side))

        # 基于 base_row_calc 计算移动颗数
        base_row_ang_center = self._cal_row_ang_center(self.__DBL_base_row)
        move_dire, move_angle = calc_move_angle(base_row_ang_center, self.__dbl_move_ang)

        # 计算row1、2、3最大移动颗数
        dbl_move_max_ang = self._cal_dbl_move_max_ang()

        # 计算row1、2、3实际移动颗数
        self.__row3_move_ang = get_each_row_move_ang(move_angle,
                                                     dbl_move_max_ang[0],
                                                     self.__dbl_move_ang,
                                                     self.__DBL_row3_enable)
        self.__row2_move_ang = get_each_row_move_ang(move_angle,
                                                     dbl_move_max_ang[1],
                                                     self.__dbl_move_ang,
                                                     self.__DBL_row2_enable)
        self.__row1_move_ang = get_each_row_move_ang(move_angle,
                                                     dbl_move_max_ang[2],
                                                     self.__dbl_move_ang,
                                                     self.__DBL_row1_enable)

        # self.__row_move_ang = [row3_move_ang, row2_move_ang, row1_move_ang]

        # label_DBL_Angle_2.setText("DBL  {}  {} LED  \r\n"
        #                           "\r\n"
        #                           "ROW3_Left_Max: {} LEDs, ROW3_Right_Max: {} LEDs\r\n"
        #                           "ROW2_Left_Max: {} LEDs, ROW2_Right_Max: {} LEDs\r\n"
        #                           "ROW1_Left_Max: {} LEDs, ROW1_Right_Max: {} LEDs\r\n"
        #                           "\r\n"
        #                           "R0W3 actual move {} LEDs\r\n"
        #                           "R0W2 actual move {} LEDs\r\n"
        #                           "R0W1 actual move {} LEDs\r\n".
        #                           format(MoveDire, MoveAngle,
        #                                  dbl_move_max_ang[0][0], dbl_move_max_ang[0][1],
        #                                  dbl_move_max_ang[1][0], dbl_move_max_ang[1][1],
        #                                  dbl_move_max_ang[2][0], dbl_move_max_ang[2][1],
        #                                  row3_move_ang, row2_move_ang, row1_move_ang))

    def _cal_row_ang_center(self, row):
        if row == DBLMoveBaseRow.row1.value:
            row_ang_center = get_row_ang_center(self.__sec45_site_value, 16)  # 翻转角度 : [12 ~~ 01 ~~ 01 ~~ 12]
        elif row == DBLMoveBaseRow.row2.value:
            row_ang_center = get_row_ang_center(self.__sec23_site_value, 16)
        elif row == DBLMoveBaseRow.row3.value:
            row_ang_center = get_row_ang_center(self.__sec01_site_value, 16)
        else:
            print(row)
            row_ang_center = 0

        return row_ang_center

    def _cal_dbl_move_max_ang(self):
        row3_dbl_l_max_ang = self.__DBL_move_para[self.__dbl_move_class_level][4]
        row3_dbl_r_max_ang = self.__DBL_move_para[self.__dbl_move_class_level][5]
        row2_dbl_l_max_ang = self.__DBL_move_para[self.__dbl_move_class_level][10]
        row2_dbl_r_max_ang = self.__DBL_move_para[self.__dbl_move_class_level][11]

        self.__row1_ang_center = self._cal_row_ang_center(DBLMoveBaseRow.row1.value)
        self.__row2_ang_center = self._cal_row_ang_center(DBLMoveBaseRow.row2.value)
        self.__row3_ang_center = self._cal_row_ang_center(DBLMoveBaseRow.row3.value)

        dire, row2_max_move_l_ang = calc_move_angle(self.__row2_ang_center,
                                                    row2_dbl_l_max_ang)
        dire, row2_max_move_r_ang = calc_move_angle(self.__row2_ang_center,
                                                    -row2_dbl_r_max_ang)
        dire, row3_max_move_l_ang = calc_move_angle(self.__row3_ang_center,
                                                    row3_dbl_l_max_ang)
        dire, row3_max_move_r_ang = calc_move_angle(self.__row3_ang_center,
                                                    -row3_dbl_r_max_ang)

        if self.__dbl_move_class_level == 7:
            row1_dbl_l_max_ang = self.__DBL_move_para[self.__dbl_move_class_level][16]
            row1_dbl_r_max_ang = self.__DBL_move_para[self.__dbl_move_class_level][17]
            dire, row1_max_move_l_ang = calc_move_angle(self.__row1_ang_center,
                                                        row1_dbl_l_max_ang)
            dire, row1_max_move_r_ang = calc_move_angle(self.__row1_ang_center,
                                                        -row1_dbl_r_max_ang)
        else:
            row1_max_move_l_ang = 0
            row1_max_move_r_ang = 0

        dbl_move_max_ang = [[row3_max_move_l_ang, row3_max_move_r_ang],
                            [row2_max_move_l_ang, row2_max_move_r_ang],
                            [row1_max_move_l_ang, row1_max_move_r_ang]]

        return dbl_move_max_ang

    def _get_para_of_dbl(self):
        # 通用参数
        row3_intensity_re = get_part_reverse_array(self.__sec01_intensity, 16)
        row2_intensity_re = get_part_reverse_array(self.__sec23_intensity, 16)
        self.__row3_notnull_intensity = get_not_null_site_int(row3_intensity_re[self.__class_level],
                                                              self.__sec01_notNull_index)
        self.__row2_notnull_intensity = get_not_null_site_int(row2_intensity_re[self.__class_level],
                                                              self.__sec23_notNull_index)

        # DBLShiftAway.Deactivated 使用参数
        dbl_min_right_ang_row3 = self.__DBL_move_para[self.__dbl_move_class_level][3]
        dbl_min_right_ang_row2 = self.__DBL_move_para[self.__dbl_move_class_level][9]
        dbl_min_left_ang_row3 = self.__DBL_move_para[self.__dbl_move_class_level][2]
        dbl_min_left_ang_row2 = self.__DBL_move_para[self.__dbl_move_class_level][8]

        self.__row3_minL_site = dbl_minleft_ang_site(self.__row3_ang_center, dbl_min_left_ang_row3)
        self.__row2_minL_site = dbl_minleft_ang_site(self.__row2_ang_center, dbl_min_left_ang_row2)

        self.__row3_minR_site = dbl_maxright_ang_site(self.__row3_ang_center, dbl_min_right_ang_row3)
        self.__row2_minR_site = dbl_maxright_ang_site(self.__row2_ang_center, dbl_min_right_ang_row2)

        # DBLShiftAway.RemainOn 使用参数
        shift_remain_ang_row3 = self.__DBL_move_para[self.__dbl_move_class_level][0]
        self.__shift_remain_int_row3 = self.__DBL_move_para[self.__dbl_move_class_level][1]
        shift_remain_ang_row2 = self.__DBL_move_para[self.__dbl_move_class_level][6]
        self.__shift_remain_int_row2 = self.__DBL_move_para[self.__dbl_move_class_level][7]

        self.__row3_remainL_site = dbl_minleft_ang_site(self.__row3_ang_center, shift_remain_ang_row3)
        self.__row3_remainR_site = dbl_maxright_ang_site(self.__row3_ang_center, -shift_remain_ang_row3)
        self.__row2_remainL_site = dbl_minleft_ang_site(self.__row2_ang_center, shift_remain_ang_row2)
        self.__row2_remainR_site = dbl_maxright_ang_site(self.__row2_ang_center, -shift_remain_ang_row2)

        # HB para
        if self.__dbl_move_class_level == 7:
            row1_intensity_re = get_part_reverse_array(self.__sec45_intensity, 16)
            self.__row1_notnull_intensity = get_not_null_site_int(row1_intensity_re[self.__class_level],
                                                                  self.__sec45_notNull_index)
            dbl_min_right_ang_row1 = self.__DBL_move_para[self.__dbl_move_class_level][15]
            dbl_min_left_ang_row1 = self.__DBL_move_para[self.__dbl_move_class_level][14]
            self.__row1_minL_site = dbl_minleft_ang_site(self.__row1_ang_center, dbl_min_left_ang_row1)
            self.__row1_minR_site = dbl_maxright_ang_site(self.__row1_ang_center, dbl_min_right_ang_row1)

            shift_remain_ang_row1 = self.__DBL_move_para[self.__dbl_move_class_level][12]
            self.__shift_remain_int_row1 = self.__DBL_move_para[self.__dbl_move_class_level][13]
            self.__row1_remainL_site = dbl_minleft_ang_site(self.__row1_ang_center, shift_remain_ang_row1)
            self.__row1_remainR_site = dbl_maxright_ang_site(self.__row1_ang_center, -shift_remain_ang_row1)

    def _cal_int_after_dbl_move(self, dbl_move_ang):
        self._get_para_of_dbl()

        if (self.__DBL_shiftAway == DBLShiftAway.RemainOn.value) and \
                (((int(float(self.__traffic_style)) == 0) and (dbl_move_ang < 0)) or
                 ((int(float(self.__traffic_style)) == 1) and (dbl_move_ang > 0))):
            # print("Remain ON")
            row3_move_intensity = remain_dbl_move_intensity(self.__row3_notnull_intensity,
                                                            self.__row3_remainL_site,
                                                            self.__row3_remainR_site,
                                                            self.__shift_remain_int_row3,
                                                            self.__row3_move_ang)
            row2_move_intensity = remain_dbl_move_intensity(self.__row2_notnull_intensity,
                                                            self.__row2_remainL_site,
                                                            self.__row2_remainR_site,
                                                            self.__shift_remain_int_row2,
                                                            self.__row2_move_ang)
            if self.__dbl_move_class_level == 7:
                row1_move_intensity = remain_dbl_move_intensity(self.__row1_notnull_intensity,
                                                                self.__row1_remainL_site,
                                                                self.__row1_remainR_site,
                                                                self.__shift_remain_int_row1,
                                                                self.__row1_move_ang)
            else:
                row1_move_intensity = []
        else:
            # print("Deac")
            row3_move_intensity = deac_dbl_move_intensity(self.__row3_notnull_intensity,
                                                          self.__row3_minR_site,
                                                          self.__row3_minL_site,
                                                          self.__row3_move_ang)
            row2_move_intensity = deac_dbl_move_intensity(self.__row2_notnull_intensity,
                                                          self.__row2_minR_site,
                                                          self.__row2_minL_site,
                                                          self.__row2_move_ang)
            if self.__dbl_move_class_level == 7:
                row1_move_intensity = deac_dbl_move_intensity(self.__row1_notnull_intensity,
                                                              self.__row1_minR_site,
                                                              self.__row1_minL_site,
                                                              self.__row1_move_ang)
            else:
                row1_move_intensity = []

        self.__row3_after_dbl_int = get_after_dbl_intensity(row3_move_intensity,
                                                            self.__sec01_notNull_index)
        self.__row2_after_dbl_int = get_after_dbl_intensity(row2_move_intensity,
                                                            self.__sec23_notNull_index)
        self.__row1_after_dbl_int = get_after_dbl_intensity(row1_move_intensity,
                                                            self.__sec45_notNull_index)

        self.__pixel_int_after_dbl = self.__row3_after_dbl_int + self.__row2_after_dbl_int + self.__row1_after_dbl_int
        self.__pixel_final_int = self.__pixel_int_after_dbl
        self._check_pixel_min_pwm()

    def _check_pixel_min_pwm(self):
        """
        The lowest pwm should be 2%.
        """
        for i in range(len(self.__pixel_final_int)):
            if self.__pixel_final_int[i] is not None:
                self.__pixel_final_int[i] = self._check_min_output_intensity(self.__pixel_final_int[i])

    @staticmethod
    def _check_min_output_intensity(input_int):
        """
        The lowest pwm should be 2%.
        @param input_int: input intensity.
        @return:
        """
        if (input_int <= 1) and (input_int != 0):
            return 2
        else:
            return input_int

    def get_drl_int(self, drl_x):
        """
        @param drl_x:
        @author:Li.Wang
        """
        if self.__side == ecu_side.left.value:
            drl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                Signal['DrlPos{}_Actv_Lf_Rq'.format(drl_x)].value)
            ti_request = self.__app.get_signal(Message.E2E_BCMtoLDCM.value,
                                               Signal.TurnLghtLeft_D_Rq.value)
        else:
            drl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                Signal['DrlPos{}_Actv_Rf_Rq'.format(drl_x)].value)
            ti_request = self.__app.get_signal(Message.E2E_BCMtoLDCM.value,
                                               Signal.TurnLghtLeft_D_Rq.value)

        drl_po_int = self.__dataset.get_drl_int_in_dataset(drl_x, drl_request, ti_request)

        return self._check_min_output_intensity(drl_po_int)

    def get_aux_int(self, aux_x):
        """
        @param drl_x:
        @author:Li.Wang
        """
        if self.__side == ecu_side.left.value:
            drl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                Signal['DrlPos{}_Actv_Lf_Rq'.format(aux_x)].value)

        else:
            drl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                Signal['DrlPos{}_Actv_Rf_Rq'.format(aux_x)].value)

        aux_int = self.__dataset.get_aux_int_in_dataset(aux_x, drl_request)

        return self._check_min_output_intensity(aux_int)

    def get_turn_indicator_int(self, function: str, drl_x: int):
        """
        @param drl_x:
        @author:Li.Wang
        """
        if self.__side == ecu_side.left.value:
            turn_indicator_request = self.__app.get_signal(Message.E2E_BCMtoLDCM.value,
                                                           Signal.TurnLghtLeft_D_Rq.value)
            if drl_x in [1, 2, 3, 4]:
                drl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                    Signal['DrlPos{}_Actv_Lf_Rq'.format(drl_x)].value)
            else:
                # 'welcome' signal
                drl_request = 4

        else:
            turn_indicator_request = self.__app.get_signal(Message.E2E_BCMtoLDCM.value,
                                                           Signal.TurnLghtRight_D_Rq.value)
            if drl_x in [1, 2, 3, 4]:
                drl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                    Signal['DrlPos{}_Actv_Rf_Rq'.format(drl_x)].value)
            else:
                # 'welcome' signal
                drl_request = 4
        # print('drl command value = ', drl_request)
        turn_indicator_int, delay_time, wait = self.__dataset.get_ti_int_in_request(turn_indicator_request, function,
                                                                                    drl_request)

        return self._check_min_output_intensity(turn_indicator_int), delay_time, wait

    def get_ezl_pixel_int(self, class_level: int):
        """
        @author:Li.Wang
        """
        self._get_class_mod_intensity()
        self.__class_level = class_level
        flb_ch, flb_int = self.get_flb_class_intensity()
        self.__pixel_final_int = self.__class_intensity[self.__class_level]
        # print('flb ', flb_ch, flb_int)
        pixel_flb_int = {}

        for key, value in flb_ch.items():
            pixel_flb_int[key] = flb_int

        return self.get_pixel_ch_int_dict(), pixel_flb_int

    def get_ezl_int(self, ezl_config: list, class_lvl: int):
        """
        @author: Li.Wang
        """
        if self.__side == ecu_side.left.value:
            ezl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                Signal.ExtLghtFront_D_RqBody.value)

        else:
            ezl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                Signal.ExtLghtFront_D_RqBody.value)
        # print('drl command value = ', drl_request)
        int_list = []
        ezl_expect_int = self.__dataset.get_ezl_int_in_dataset(ezl_config, ezl_request, 0, 0)

        # pixel intensity handler
        pixel_int, flb_int = self.get_ezl_pixel_int(class_lvl)
        # print('ezl_request = ', ezl_request)

        if ezl_request == 2 and ezl_config[3] == 1:
            for keys, values in pixel_int.items():
                # print('ezl_config = ', ezl_config[index])
                for key, value in values.items():
                    int_list.append(self._check_min_output_intensity(values.get(key)))
                pixel_int[keys] = int_list
                int_list = []
        else:
            for keys, values in pixel_int.items():
                for key, value in values.items():
                    int_list.append(0)
                pixel_int[keys] = int_list
                int_list = []

        if ezl_request == 2 and ezl_config[3] == 1 and ezl_config[9] == 1:
            for key, value in flb_int.items():
                pixel_int[key] = value
        else:
            for key, value in flb_int.items():
                pixel_int[key] = 0

        for key, value in ezl_expect_int.items():
            ezl_expect_int[key] = self._check_min_output_intensity(value)

        return ezl_expect_int, pixel_int

    def get_po_interrupt_ezl_int(self, ezl_config: list, drl_x: int):
        """
        @author: Li.Wang
        """
        if self.__side == ecu_side.left.value:
            ezl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                Signal.ExtLghtFront_D_RqBody.value)
            if drl_x in [1, 2, 3, 4]:
                drl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                    Signal['DrlPos{}_Actv_Lf_Rq'.format(drl_x)].value)

        else:
            ezl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                Signal.ExtLghtFront_D_RqBody.value)
            if drl_x in [1, 2, 3, 4]:
                drl_request = self.__app.get_signal(Message.BaseFeaturesActvRq.value,
                                                    Signal['DrlPos{}_Actv_Rf_Rq'.format(drl_x)].value)

        # print('drl command value = ', drl_request)
        ezl_expect_int = self.__dataset.get_ezl_int_in_dataset(ezl_config, ezl_request, drl_request, drl_x)
        # print('ezl_expect_int = ', ezl_expect_int)
        for key, value in ezl_expect_int.items():
            ezl_expect_int[key] = self._check_min_output_intensity(value)

        return ezl_expect_int

    def get_iocontrol_ezl_int(self, fun: str, iocontrol_request: int):
        """
        @author: Li.Wang
        """
        # print('drl command value = ', drl_request)
        ezl_expect_int = self.__dataset.get_ezl_iocontrol_int_in_dataset(fun, iocontrol_request)
        # print('ezl_expect_int = ', ezl_expect_int)
        for key, value in ezl_expect_int.items():
            ezl_expect_int[key] = self._check_min_output_intensity(value)
        # print('fun = ', fun)
        # print(ezl_expect_int)
        int_list = []
        pixel_int = {}
        if iocontrol_request == 1:
            if fun == 'lb':
                pixel_int, flb_int = self.get_ezl_pixel_int(0)
                for key, value in pixel_int.items():
                    for keys, values in value.items():
                        int_list.append(self._check_min_output_intensity(value.get(keys)))
                    pixel_int[key] = int_list
                    int_list = []
                for key, value in flb_int.items():
                    pixel_int[key] = value
            elif fun == 'hb':
                print('hb')
                pixel_int, flb_int = self.get_ezl_pixel_int(6)
                for key, value in pixel_int.items():
                    for keys, values in value.items():
                        int_list.append(self._check_min_output_intensity(value.get(keys)))
                    pixel_int[key] = int_list
                    int_list = []
                for key, value in flb_int.items():
                    pixel_int[key] = value

        print('pixel_int = ', pixel_int)
        return ezl_expect_int, pixel_int

    def get_e2e_int(self, dtc_expect: list):
        """
        @author: Li.Wang
        """
        int_list = []
        pixel_int = {}
        if self.__side == ecu_side.left.value:
            lb_request = self.__app.get_signal(Message.E2E_BCMtoLDCM.value,
                                                Signal.HeadLampLoFlOn_B_Stat.value)

        else:
            lb_request = self.__app.get_signal(Message.E2E_BCMtoLDCM.value,
                                                Signal.HeadLampLoFrOn_B_Stat.value)

        # print('ezl_expect_int = ', ezl_expect_int)
        ezl_expect_int = self.__dataset.get_e2e_int(dtc_expect)
        for key, value in ezl_expect_int.items():
            ezl_expect_int[key] = self._check_min_output_intensity(value)

        if (dtc_expect != [] and dtc_expect != [1]) or lb_request == 1:
            pixel_int, flb_int = self.get_ezl_pixel_int(0)
            for key, value in pixel_int.items():
                for keys, values in value.items():
                    int_list.append(self._check_min_output_intensity(value.get(keys)))
                pixel_int[key] = int_list
                int_list = []
            for key, value in flb_int.items():
                pixel_int[key] = value

        # print('pixel_int = ', pixel_int)
        return ezl_expect_int, pixel_int
